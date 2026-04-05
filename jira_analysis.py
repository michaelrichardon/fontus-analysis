"""
Jira Fontus – Bug-Analyse & Statuswechsel-Auswertung
=====================================================
Schritt 1: Alle Bugs aus dem Projekt Fontus abrufen
Schritt 2: Zeiten zwischen Statuswechseln je Bug berechnen
"""

import os
import requests
from datetime import datetime, timezone
import json
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Konfiguration – Werte in .env oder direkt hier eintragen
# ---------------------------------------------------------------------------
JIRA_BASE_URL      = os.getenv("JIRA_BASE_URL",          "https://jira.dein-unternehmen.de")
JIRA_PAT           = os.getenv("JIRA_PAT",                "DEIN_PERSONAL_ACCESS_TOKEN")
PROJECT_KEY        = os.getenv("JIRA_PROJECT",            "FONTUS")
# Wird beim ersten Start automatisch ermittelt – oder hier manuell eintragen, z. B. "customfield_11101"
FEATURE_TEAM_FIELD = os.getenv("JIRA_FEATURE_TEAM_FIELD", "")
# Pfad zur SQLite-Datenbank (wird angelegt falls nicht vorhanden)
DB_PATH            = os.getenv("JIRA_DB_PATH",            "fontus_analysis.db")

# ---------------------------------------------------------------------------
# Feature Teams – welche Teams werden abgefragt
# ---------------------------------------------------------------------------
FEATURE_TEAMS = [
    "DL-01",
    "DL-02",
    "DL-GenAI",
    "DLH",
    "DLH Source",
    "DL-Output",
    "UI&S",
    "User Interfaces",
    "T2-Team",
    "Rule Engine",
    "Rule Validation",
]

# Mapping: Org-Team → Aggregierter Name (für Auswertung & Export)
TEAM_MAPPING = {
    "DL-01":           "DLH",
    "DL-02":           "DLH",
    "DL-GenAI":        "DLH",
    "DLH":             "DLH",
    "DLH Source":      "DLH",
    "DL-Output":       "DLH",
    "UI&S":            "User Interfaces",
    "User Interfaces": "User Interfaces",
    "T2-Team":         "Rule Engine",
    "Rule Engine":     "Rule Engine",
    "Rule Validation": "Rule Validation",
}

# ---------------------------------------------------------------------------
# Label-Konfiguration
# ---------------------------------------------------------------------------

# Issues mit mindestens einem dieser Labels werden komplett ausgeschlossen (JQL NOT IN)
# Groß-/Kleinschreibung wird beim JQL-Filter von Jira beachtet – alle Varianten eintragen
EXCLUDE_LABELS = [
    "Not-Testable",
    "non-testable",
]

# Labels die als eigene Spalte (True / leer) erscheinen sollen
# Format: { "Jira-Label": "Spaltenname" }
LABEL_COLUMNS = {
    "E2ESrcWatch": "E2ESrcWatch",
    "MigWatch":    "MigWatch",
}

# ---------------------------------------------------------------------------
# Status-Filter – kein Filter, alle Bugs werden geladen.
# Die Unterscheidung offen/geschlossen erfolgt in den Reports und im Export
# über die Spalte "Abgeschlossen" (True/False).
# ---------------------------------------------------------------------------
STATUS_FILTER = []   # leer = alle Status

# Status die als "abgeschlossen" gelten (für die Spalte "Abgeschlossen")
CLOSED_STATUSES   = {"Done", "Closed"}
# Status die als "rejected" gelten – eigene Kennung in der Spalte "Abgeschlossen"
REJECTED_STATUSES = {"Rejected"}

# ---------------------------------------------------------------------------
# Zusätzliche Custom Fields
# ---------------------------------------------------------------------------
SOURCE_FIELD     = "customfield_11104"   # Quelle
ROOT_CAUSE_FIELD = "customfield_11236"   # Root Cause
TEST_LEVEL_FIELD = "customfield_11106"   # Test Level
COMPONENT_FIELD  = "components"          # Komponenten (Mehrfachauswahl, Standard-Jira-Feld)

# Labels die als Release-Werte interpretiert werden.
# Alle passenden Labels eines Issues werden kommagetrennt in "Releases" eingetragen.
RELEASE_LABELS = [
    "R2.1", "R2.2", "R2.3",
    "R3.1", "R3.2", "R3.3", "R3.4", "R3.5", "R3.6", "R3.7", "R3.8",
    "R3.9", "R3.10", "R3.11", "R3.12", "R3.13", "R3.14", "R3.15", "R3.16", "R3.17",
]

# ---------------------------------------------------------------------------
# Retest-Zählung
# Status dessen Eintritte als "Retest" gezählt werden.
# Segmente mit Dauer = 0 Sekunden (Fehlbedienungen) werden vorher eliminiert.
# ---------------------------------------------------------------------------
RETEST_STATUS = "Ready for Test"

# ---------------------------------------------------------------------------
# Status → Kategorie Mapping für die aggregierte Liegezeit-Auswertung
# ---------------------------------------------------------------------------
STATUS_CATEGORY_MAPPING = {
    "Stopped":               "BLOCKED",
    "Blocked":               "BLOCKED",
    "To Do":                 "Development",
    "In Progress":           "Development",
    "In Review":             "Development",
    "Reopened":              "Development",
    "Ready For Deployment":  "Ready For Deployment",
    "Ready for Test":        "Test",
    "In Test":               "Test",
}

# ---------------------------------------------------------------------------
# HTTP-Session mit Bearer Token (Personal Access Token)
# ---------------------------------------------------------------------------
session = requests.Session()
session.headers = {
    "Authorization": f"Bearer {JIRA_PAT}",
    "Accept":        "application/json",
    "Content-Type":  "application/json",
}

# SSL-Zertifikat-Prüfung
# True        = Zertifikat prüfen (Standard, empfohlen in Produktion)
# False       = Prüfung deaktivieren (schnell, nur zum Testen!)
# "/pfad/zum/ca-bundle.crt" = eigenes CA-Zertifikat verwenden (empfohlen)
SSL_VERIFY = os.getenv("JIRA_SSL_VERIFY", "false").lower()
if SSL_VERIFY == "false":
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    session.verify = False
elif SSL_VERIFY == "true":
    session.verify = True
else:
    session.verify = SSL_VERIFY  # Pfad zum CA-Bundle

def get(path: str, params: dict = None) -> dict:
    """GET-Wrapper mit Fehlerbehandlung."""
    url = f"{JIRA_BASE_URL}/rest/api/2/{path}"
    resp = session.get(url, params=params)
    resp.raise_for_status()
    return resp.json()


# ===========================================================================
# FELDERKENNUNG – Findet das technische customfield für "Feature Team"
# ===========================================================================

def discover_feature_team_field(search_term: str = "feature team") -> str | None:
    """
    Durchsucht alle Felder der Jira-Instanz nach dem Namen 'Feature Team'.
    Gibt den technischen Feldnamen zurück (z. B. 'customfield_10204').
    Wenn mehrere Treffer, werden alle zur Auswahl angezeigt.
    """
    print(f"\n[Felderkennung] Suche nach Feld '{search_term}' in allen Custom Fields ...")
    fields = get("field")

    matches = [
        f for f in fields
        if search_term.lower() in f.get("name", "").lower()
    ]

    if not matches:
        print(f"  ✗ Kein Feld mit '{search_term}' gefunden.\n")
        print("  Alle verfügbaren Custom Fields:")
        custom = [f for f in fields if f["id"].startswith("customfield_")]
        for f in sorted(custom, key=lambda x: x["name"]):
            print(f"    {f['id']:<30} → {f['name']}")
        return None

    if len(matches) == 1:
        field_id = matches[0]["id"]
        print(f"  ✓ Gefunden: '{matches[0]['name']}' → {field_id}\n")
        return field_id

    # Mehrere Treffer
    print(f"  Mehrere Treffer – bitte in .env eintragen:")
    for f in matches:
        print(f"    {f['id']:<30} → {f['name']}")
    print()
    return matches[0]["id"]   # Ersten Treffer als Standard nehmen


# ===========================================================================
# SCHRITT 1 – Alle Bugs (issuetype = Bug) aus Projekt Fontus abrufen
# ===========================================================================

def fetch_bugs(project: str = PROJECT_KEY,
               feature_team_field: str = "",
               feature_teams: list[str] = None,
               exclude_labels: list[str] = None,
               status_filter: list[str] = None) -> list[dict]:
    """
    Gibt alle Bugs zurück, gefiltert auf die definierten Feature Teams und Status.
    Issues mit einem der exclude_labels werden via JQL ausgeschlossen.
    Nutzt JQL IN-Operator für alle Teams und Status in einem einzigen Query.
    """
    # Team-Filter
    team_filter = ""
    if feature_team_field and feature_teams:
        cf_number   = feature_team_field.replace("customfield_", "")
        teams_jql   = ", ".join(f'"{t}"' for t in feature_teams)
        team_filter = f" AND cf[{cf_number}] in ({teams_jql})"

    # Label-Ausschluss: labels NOT IN ("Not-Testable", "non-testable")
    label_filter = ""
    if exclude_labels:
        labels_jql   = ", ".join(f'"{l}"' for l in exclude_labels)
        label_filter = f" AND labels not in ({labels_jql})"

    # Status-Filter: status IN ("Done", "Closed")
    status_jql_filter = ""
    if status_filter:
        status_jql  = ", ".join(f'"{s}"' for s in status_filter)
        status_jql_filter = f" AND status in ({status_jql})"

    jql = (
        f'project = "{project}" AND issuetype = Bug'
        f'{team_filter}'
        f'{label_filter}'
        f'{status_jql_filter}'
        f' ORDER BY created ASC'
    )

    start_at    = 0
    max_results = 100
    bugs        = []

    team_info   = f"{len(feature_teams)} Teams" if feature_teams else "kein Team-Filter"
    excl_info   = f", ausgeschlossen Labels: {', '.join(exclude_labels)}" if exclude_labels else ""
    status_info = f", Status: {', '.join(status_filter)}" if status_filter else ", alle Status"
    print(f"\n[Schritt 1] Lade Bugs für Projekt '{project}' (Filter: {team_info}{excl_info}{status_info}) ...")
    if feature_teams:
        print(f"  Teams: {', '.join(feature_teams)}")

    while True:
        data = get("search", params={
            "jql":        jql,
            "startAt":    start_at,
            "maxResults": max_results,
            "fields":     f"summary,status,priority,assignee,reporter,created,updated,resolutiondate,labels,components,{SOURCE_FIELD},{ROOT_CAUSE_FIELD},{TEST_LEVEL_FIELD}{(',' + feature_team_field) if feature_team_field else ''}",
        })

        issues = data.get("issues", [])
        bugs.extend(issues)
        print(f"  → {len(bugs)} / {data['total']} Issues geladen")

        if start_at + max_results >= data["total"]:
            break
        start_at += max_results

    print(f"[Schritt 1] Fertig – {len(bugs)} Bugs gefunden.\n")
    return bugs


def get_team_value(issue: dict, field_id: str) -> str:
    """Liest den Feature-Team-Wert aus einem Issue (kann String oder Objekt sein)."""
    raw = issue["fields"].get(field_id)
    if raw is None:
        return "–"
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or str(raw)
    return str(raw)


def get_labels(issue: dict) -> list[str]:
    """Gibt die Labels eines Issues als Liste zurück."""
    return issue["fields"].get("labels") or []


def get_source(issue: dict) -> str:
    """Liest den Wert von SOURCE_FIELD (Quelle) aus einem Issue."""
    raw = issue["fields"].get(SOURCE_FIELD)
    if raw is None:
        return ""
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or str(raw)
    return str(raw)


def get_root_cause(issue: dict) -> str:
    """Liest den Wert von ROOT_CAUSE_FIELD aus einem Issue."""
    raw = issue["fields"].get(ROOT_CAUSE_FIELD)
    if raw is None:
        return ""
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or str(raw)
    if isinstance(raw, list):
        return ", ".join(
            (v.get("value") or v.get("name") or str(v)) if isinstance(v, dict) else str(v)
            for v in raw
        )
    return str(raw)


def get_test_level(issue: dict) -> str:
    """Liest den Wert von TEST_LEVEL_FIELD aus einem Issue."""
    raw = issue["fields"].get(TEST_LEVEL_FIELD)
    if raw is None:
        return ""
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or str(raw)
    if isinstance(raw, list):
        return ", ".join(
            (v.get("value") or v.get("name") or str(v)) if isinstance(v, dict) else str(v)
            for v in raw
        )
    return str(raw)


def get_components(issue: dict) -> str:
    """Liest die Komponenten (Mehrfachauswahl) aus einem Issue, kommagetrennt."""
    raw = issue["fields"].get("components") or []
    return ", ".join(c.get("name", str(c)) for c in raw if isinstance(c, dict))


def get_releases(labels: list[str]) -> str:
    """Filtert Release-Labels aus der Label-Liste und gibt sie kommagetrennt zurück."""
    found = [lbl for lbl in labels if lbl in RELEASE_LABELS]
    return ", ".join(sorted(found))


def print_bug_summary(bugs: list[dict], field_id: str = "") -> None:
    """Gibt eine kompakte Übersicht aller Bugs aus."""
    label_headers = "".join(f"{col:<10}" for col in LABEL_COLUMNS.values())
    print(f"{'KEY':<16} {'TEAM':<18} {'AGG':<18} {'STATUS':<18} {'PRIO':<8} {'RETEST':>6}  {'QUELLE':<16} {'RELEASES':<18} {label_headers}{'ZUSAMMENFASSUNG'}")
    print("-" * (130 + 10 * len(LABEL_COLUMNS)))
    for issue in bugs:
        f        = issue["fields"]
        key      = issue["key"]
        status   = f["status"]["name"]
        priority = (f.get("priority") or {}).get("name", "")
        summary  = f["summary"][:40]
        team     = get_team_value(issue, field_id) if field_id else "–"
        agg      = TEAM_MAPPING.get(team, team)
        labels   = get_labels(issue)
        source   = get_source(issue)
        releases = get_releases(labels)
        label_vals = "".join(
            f"{'True':<10}" if lbl in labels else f"{'': <10}"
            for lbl in LABEL_COLUMNS.keys()
        )
        # Retest hier noch nicht berechenbar (kein Changelog in Schritt 1) → Platzhalter
        print(f"{key:<16} {team:<18} {agg:<18} {status:<18} {priority:<8} {'?':>6}  {source:<16} {releases:<18} {label_vals}{summary}")


# ===========================================================================
# SCHRITT 2 – Statuswechsel-Zeiten je Bug auswerten
# ===========================================================================

def fetch_changelog(issue_key: str) -> list[dict]:
    """
    Liest den vollständigen Changelog eines Issues.
    Jira Server/DC: changelog ist Teil des Issue-Endpunkts (expand=changelog),
    NICHT ein separater /changelog Endpunkt (das ist nur Jira Cloud).
    Gibt nur Status-Übergänge zurück.
    """
    transitions = []
    start_at    = 0
    max_results = 100

    while True:
        data = get(
            f"issue/{issue_key}",
            params={
                "expand":     "changelog",
                "startAt":    start_at,
                "maxResults": max_results,
            },
        )
        changelog  = data.get("changelog", {})
        histories  = changelog.get("histories", [])

        for history in histories:
            created = parse_dt(history["created"])
            for item in history.get("items", []):
                if item["field"] == "status":
                    transitions.append({
                        "timestamp": created,
                        "from":      item.get("fromString", "–"),
                        "to":        item.get("toString",   "–"),
                        "author":    history.get("author", {}).get("displayName", "unbekannt"),
                    })

        total = changelog.get("total", len(histories))
        if start_at + max_results >= total:
            break
        start_at += max_results

    transitions.sort(key=lambda x: x["timestamp"])
    return transitions


def parse_dt(ts: str) -> datetime:
    """Parst einen Jira-Zeitstempel zu einem timezone-aware datetime."""
    # Jira liefert z. B. "2024-03-15T09:23:41.000+0200"
    ts = ts[:26] + ts[26:].replace(":", "")   # +02:00 → +0200
    try:
        return datetime.strptime(ts, "%Y-%m-%dT%H:%M:%S.%f%z")
    except ValueError:
        return datetime.strptime(ts[:19], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=timezone.utc)


def fmt_duration(seconds: float) -> str:
    """Formatiert Sekunden lesbar als 'd h m'."""
    total_minutes = int(seconds // 60)
    days,  rem    = divmod(total_minutes, 1440)
    hours, mins   = divmod(rem, 60)
    parts = []
    if days:  parts.append(f"{days}d")
    if hours: parts.append(f"{hours}h")
    parts.append(f"{mins}m")
    return " ".join(parts) or "0m"


def analyse_transitions(bugs: list[dict], field_id: str = "") -> list[dict]:
    """
    Berechnet für jeden Bug die Verweildauer in jedem Status.
    Reichert jeden Bug mit Team und aggregiertem Team-Namen an.
    """
    print("[Schritt 2] Werte Statuswechsel aus ...\n")
    results = []

    for issue in bugs:
        key     = issue["key"]
        summary = issue["fields"]["summary"][:55]
        created = parse_dt(issue["fields"]["created"])
        team    = get_team_value(issue, field_id) if field_id else "–"
        agg     = TEAM_MAPPING.get(team, team)
        labels  = get_labels(issue)
        label_vals = {col: ("True" if lbl in labels else "") for lbl, col in LABEL_COLUMNS.items()}
        priority = (issue["fields"].get("priority") or {}).get("name", "")
        source     = get_source(issue)
        root_cause = get_root_cause(issue)
        test_level = get_test_level(issue)
        components = get_components(issue)
        releases   = get_releases(labels)

        transitions = fetch_changelog(key)

        # Einstiegspunkt: Erstellung des Issues
        segments = []
        prev_ts     = created
        prev_status = issue["fields"]["status"]["name"]  # aktueller Status als Fallback

        # Statuswechsel rückwärts aufbauen: erster Übergang zeigt den Anfangsstatus
        if transitions:
            prev_status = transitions[0]["from"]

        for t in transitions:
            duration_sec = (t["timestamp"] - prev_ts).total_seconds()
            segments.append({
                "status":       prev_status,
                "entered_at":   prev_ts.strftime("%Y-%m-%d %H:%M"),
                "left_at":      t["timestamp"].strftime("%Y-%m-%d %H:%M"),
                "duration_sec": duration_sec,
                "duration_fmt": fmt_duration(duration_sec),
            })
            prev_ts     = t["timestamp"]
            prev_status = t["to"]

        # Letzter (offener) Status – bis jetzt
        now = datetime.now(timezone.utc)
        segments.append({
            "status":       prev_status,
            "entered_at":   prev_ts.strftime("%Y-%m-%d %H:%M"),
            "left_at":      "(offen)",
            "duration_sec": (now - prev_ts).total_seconds(),
            "duration_fmt": fmt_duration((now - prev_ts).total_seconds()) + " (laufend)",
        })

        # ── Fehlbedienungen eliminieren ──────────────────────────────────
        # Segmente mit Dauer 0 Sekunden sind Durchklick-Korrekturen und
        # verfälschen die Retest-Zählung → werden entfernt.
        clean_segments = [s for s in segments if s["duration_sec"] > 0]

        # ── Retest zählen ────────────────────────────────────────────────
        # Jedes echte Eintreten in RETEST_STATUS = 1 Retest-Durchlauf.
        # Der erste Eintritt zählt nicht (erster Test, kein RE-test).
        retest_count = max(0, sum(1 for s in clean_segments if s["status"] == RETEST_STATUS) - 1)

        # ── Offen / Geschlossen / Rejected ──────────────────────────────
        current_status = issue["fields"]["status"]["name"]
        if current_status in CLOSED_STATUSES:
            closed_label = "Ja"
        elif current_status in REJECTED_STATUSES:
            closed_label = "reject"
        else:
            closed_label = "Nein"

        results.append({
            "key":          key,
            "summary":      summary,
            "team":         team,
            "team_group":   agg,
            "priority":     priority,
            "source":       source,
            "root_cause":   root_cause,
            "test_level":   test_level,
            "components":   components,
            "releases":     releases,
            "retest_count": retest_count,
            "is_closed":    closed_label,
            "status":       current_status,
            **label_vals,
            "segments":     clean_segments,
        })

    print(f"[Schritt 2] Fertig – {len(results)} Bugs ausgewertet.\n")
    return results


def print_transition_report(results: list[dict]) -> None:
    """Druckt einen lesbaren Statuswechsel-Bericht."""
    print("=" * 90)
    print("STATUSWECHSEL-REPORT – Projekt Fontus / Bugs")
    print("=" * 90)

    for r in results:
        team_info    = f"{r['team']} → {r['team_group']}" if r['team'] != r['team_group'] else r['team']
        active_flags = [col for col in LABEL_COLUMNS.values() if r.get(col)]
        flags_str    = f"  [{', '.join(active_flags)}]" if active_flags else ""
        releases_str = f"  Releases: {r['releases']}" if r['releases'] else ""
        retest_str   = f"  Retests: {r['retest_count']}"
        closed_str   = {"Ja": "✓ Geschlossen", "reject": "✗ Rejected", "Nein": "⏳ Offen"}.get(r["is_closed"], r["is_closed"])
        print(f"\n▶ {r['key']}  [{team_info}]  Prio: {r['priority']}  {closed_str}  Quelle: {r['source']}{releases_str}{retest_str}{flags_str}")
        print(f"  {r['summary']}")
        print(f"  {'STATUS':<25} {'EINGETRETEN':<18} {'VERLASSEN':<18} {'DAUER'}")
        print(f"  {'-'*80}")
        for s in r["segments"]:
            print(f"  {s['status']:<25} {s['entered_at']:<18} {s['left_at']:<18} {s['duration_fmt']}")


def export_json(results: list[dict], path: str = "fontus_transitions.json") -> None:
    """Speichert die Ergebnisse als JSON-Datei."""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    print(f"\n[Export JSON] Ergebnisse gespeichert: {path}")


def export_excel(results: list[dict], path: str = "fontus_analysis.xlsx") -> None:
    """
    Exportiert alle Auswertungen in ein Excel-File mit folgenden Tabs:
      1. Defects        – Alle Bugs mit Stammdaten
      2. Statuswechsel  – Jeder Status-Segment je Bug (flach)
      3. Retest-Matrix  – Defect-ID, Team, Prio, Retest-Anzahl
      4. Liegezeiten    – Aggregiert nach Team × Kategorie × Prio
      5. Root Cause     – Häufigkeit je Root Cause (nur Done/Closed, ohne leere)
      6. Komponenten    – Nennungen & betroffene Defects je Komponente
      7. Charts         – Grafische Darstellungen
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from collections import defaultdict

    PRIO_ORDER     = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    CATEGORY_ORDER = {"BLOCKED": 0, "Development": 1, "Test": 2, "Ready For Deployment": 3, "Sonstige": 99}

    HDR_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ALT_FILL  = PatternFill("solid", start_color="DEEAF1", end_color="DEEAF1")
    BODY_FONT = Font(name="Arial", size=10)
    BOLD_FONT = Font(name="Arial", bold=True, size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")
    PRIO_COLORS = {"Critical": "C00000", "High": "FF0000", "Medium": "ED7D31", "Low": "70AD47"}

    def hdr(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CENTER

    def cel(ws, row, col, value, bold=False, center=False, fill=None, num_fmt=None, color=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, size=10,
                      color=("FFFFFF" if color else "000000"))
        c.alignment = CENTER if center else LEFT
        if fill:    c.fill = fill
        if color:   c.fill = PatternFill("solid", start_color=color, end_color=color)
        if num_fmt: c.number_format = num_fmt

    def widths(ws, w: dict):
        for k, v in w.items():
            ws.column_dimensions[k].width = v

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── TAB 1: DEFECTS ────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Defects")
    label_cols = list(LABEL_COLUMNS.values())
    h1 = ["Defect-ID", "Zusammenfassung", "Team", "Team (Agg.)", "Status", "Abgeschlossen",
          "Priorität", "Quelle", "Root Cause", "Test Level", "Component", "Releases", "Retest-Anzahl"] + label_cols
    for ci, h in enumerate(h1, 1):
        hdr(ws1, 1, ci, h)

    for ri, r in enumerate(sorted(results, key=lambda x: (x["team_group"], PRIO_ORDER.get(x["priority"], 99))), 2):
        f = ALT_FILL if ri % 2 == 0 else None
        pc = PRIO_COLORS.get(r["priority"])
        cel(ws1, ri,  1, r["key"],          bold=True, fill=f)
        cel(ws1, ri,  2, r["summary"],      fill=f)
        cel(ws1, ri,  3, r["team"],         fill=f)
        cel(ws1, ri,  4, r["team_group"],   fill=f)
        cel(ws1, ri,  5, r["status"],       fill=f)
        cel(ws1, ri,  6, r["is_closed"],    fill=f, center=True)
        cel(ws1, ri,  7, r["priority"],     bold=bool(pc), color=pc)
        cel(ws1, ri,  8, r["source"],       fill=f)
        cel(ws1, ri,  9, r["root_cause"],   fill=f)
        cel(ws1, ri, 10, r["test_level"],   fill=f)
        cel(ws1, ri, 11, r["components"],   fill=f)
        cel(ws1, ri, 12, r["releases"],     fill=f)
        cel(ws1, ri, 13, r["retest_count"], fill=f, center=True)
        for ci, col in enumerate(label_cols, 14):
            cel(ws1, ri, ci, r.get(col, ""), fill=f, center=True)

    widths(ws1, {"A": 16, "B": 55, "C": 18, "D": 18, "E": 20, "F": 14,
                 "G": 10, "H": 18, "I": 22, "J": 18, "K": 28, "L": 18, "M": 12})
    ws1.freeze_panes = "B2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}1"

    # ── TAB 2: STATUSWECHSEL ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Statuswechsel")
    h2 = ["Defect-ID", "Team (Agg.)", "Priorität", "Abgeschlossen", "Status",
          "Eingetreten", "Verlassen", "Dauer (h)", "Kategorie"]
    for ci, h in enumerate(h2, 1):
        hdr(ws2, 1, ci, h)

    r2 = 2
    for r in sorted(results, key=lambda x: x["key"]):
        for seg in r["segments"]:
            f     = ALT_FILL if r2 % 2 == 0 else None
            dur_h = round(seg["duration_sec"] / 3600, 2)
            cat   = STATUS_CATEGORY_MAPPING.get(seg["status"], "Sonstige")
            cel(ws2, r2, 1, r["key"],                            bold=True, fill=f)
            cel(ws2, r2, 2, r["team_group"],                     fill=f)
            cel(ws2, r2, 3, r["priority"],                       fill=f)
            cel(ws2, r2, 4, r["is_closed"],  fill=f, center=True)
            cel(ws2, r2, 5, seg["status"],                       fill=f)
            cel(ws2, r2, 6, seg["entered_at"],                   fill=f, center=True)
            cel(ws2, r2, 7, seg["left_at"],                      fill=f, center=True)
            cel(ws2, r2, 8, dur_h,                               fill=f, center=True, num_fmt="#,##0.00")
            cel(ws2, r2, 9, cat,                                 fill=f)
            r2 += 1

    widths(ws2, {"A": 16, "B": 18, "C": 10, "D": 14, "E": 25,
                 "F": 18, "G": 18, "H": 12, "I": 22})
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(h2))}1"

    # ── TAB 3: RETEST-MATRIX ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Retest-Matrix")
    h3 = ["Defect-ID", "Team (Agg.)", "Priorität", "Abgeschlossen", "Retest-Anzahl"]
    for ci, h in enumerate(h3, 1):
        hdr(ws3, 1, ci, h)

    for ri, r in enumerate(sorted(results, key=lambda x: (x["team_group"], PRIO_ORDER.get(x["priority"], 99), -x["retest_count"])), 2):
        f  = ALT_FILL if ri % 2 == 0 else None
        pc = PRIO_COLORS.get(r["priority"])
        cel(ws3, ri, 1, r["key"],                           bold=True, fill=f)
        cel(ws3, ri, 2, r["team_group"],                    fill=f)
        cel(ws3, ri, 3, r["priority"],                      bold=bool(pc), color=pc)
        cel(ws3, ri, 4, r["is_closed"], fill=f, center=True)
        cel(ws3, ri, 5, r["retest_count"],                  fill=f, center=True)

    widths(ws3, {"A": 16, "B": 20, "C": 12, "D": 14, "E": 14})
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = "A1:E1"

    # ── TAB 4: LIEGEZEITEN ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Liegezeiten")
    h4 = ["Team (Agg.)", "Kategorie", "Priorität", "Abgeschlossen",
          "Anzahl Defects", "∅ Liegezeit (h)", "Gesamt (h)"]
    for ci, h in enumerate(h4, 1):
        hdr(ws4, 1, ci, h)

    agg4 = defaultdict(lambda: {"total_sec": 0.0, "defect_keys": set()})
    for r in results:
        for seg in r["segments"]:
            if seg["duration_sec"] <= 0:
                continue
            cat = STATUS_CATEGORY_MAPPING.get(seg["status"], "Sonstige")
            agg4[(r["team_group"], cat, r["priority"], r["is_closed"])]["total_sec"]    += seg["duration_sec"]
            agg4[(r["team_group"], cat, r["priority"], r["is_closed"])]["defect_keys"].add(r["key"])

    sorted_k4 = sorted(agg4.keys(),
        key=lambda k: (k[0], CATEGORY_ORDER.get(k[1], 99), PRIO_ORDER.get(k[2], 99), {"Ja": 0, "Nein": 1, "reject": 2}.get(k[3], 9)))

    ri4 = 2; prev_grp = None
    for k in sorted_k4:
        group, cat, prio, closed = k
        data    = agg4[k]
        cnt     = len(data["defect_keys"])
        total_h = round(data["total_sec"] / 3600, 2)
        avg_h   = round(total_h / cnt, 2) if cnt else 0
        if group != prev_grp and prev_grp is not None:
            ri4 += 1
        prev_grp = group
        f = ALT_FILL if ri4 % 2 == 0 else None
        cel(ws4, ri4, 1, group,                             bold=True, fill=f)
        cel(ws4, ri4, 2, cat,                               fill=f)
        cel(ws4, ri4, 3, prio,                              fill=f)
        cel(ws4, ri4, 4, closed,        fill=f, center=True)
        cel(ws4, ri4, 5, cnt,                               fill=f, center=True)
        cel(ws4, ri4, 6, avg_h,                             fill=f, center=True, num_fmt="#,##0.00")
        cel(ws4, ri4, 7, total_h,                           fill=f, center=True, num_fmt="#,##0.00")
        ri4 += 1

    widths(ws4, {"A": 22, "B": 22, "C": 12, "D": 14, "E": 16, "F": 18, "G": 16})
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = "A1:G1"

    # ── TAB 5: ROOT CAUSE ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("Root Cause")

    # Nur abgeschlossene Bugs (Ja) mit befülltem Root Cause
    closed_results    = [r for r in results if r["is_closed"] == "Ja"]
    rc_results        = [r for r in closed_results if r["root_cause"]]
    rc_empty_count    = len(closed_results) - len(rc_results)

    from collections import Counter
    rc_counter: Counter = Counter()
    for r in rc_results:
        rc_counter[r["root_cause"]] += 1

    total_closed  = len(closed_results)
    total_with_rc = len(rc_results)

    # Header
    rc_headers = ["Root Cause", "Anzahl", "Anteil %"]
    for ci, h in enumerate(rc_headers, 1):
        hdr(ws5, 1, ci, h)

    # Zeile 2: Gesamt-Info inkl. Hinweis auf ausgefilterte leere Einträge
    info_text = (
        f"Basis: {total_with_rc} abgeschlossene Bugs mit Root Cause "
        f"(von {total_closed} gesamt – {rc_empty_count} ohne Angabe ausgeblendet)"
    )
    ws5.cell(row=2, column=1, value=info_text)
    ws5.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=9, color="595959")
    ws5.merge_cells("A2:C2")

    # Daten: sortiert nach Häufigkeit absteigend
    ri5 = 3
    for rc, cnt in rc_counter.most_common():
        pct = round(cnt / total_with_rc * 100, 1) if total_with_rc else 0
        f   = ALT_FILL if ri5 % 2 == 0 else None
        cel(ws5, ri5, 1, rc,  fill=f)
        cel(ws5, ri5, 2, cnt, fill=f, center=True)
        ws5.cell(row=ri5, column=3).value          = pct / 100
        ws5.cell(row=ri5, column=3).number_format  = "0.0%"
        ws5.cell(row=ri5, column=3).alignment      = CENTER
        ws5.cell(row=ri5, column=3).font           = BODY_FONT
        if f:
            ws5.cell(row=ri5, column=3).fill = f
        ri5 += 1

    # Summenzeile
    ws5.cell(row=ri5, column=1, value="Gesamt").font = BOLD_FONT
    ws5.cell(row=ri5, column=2, value=total_with_rc).font = BOLD_FONT
    ws5.cell(row=ri5, column=3, value=1.0).font          = BOLD_FONT
    ws5.cell(row=ri5, column=3).number_format             = "0.0%"

    widths(ws5, {"A": 35, "B": 12, "C": 12})
    ws5.freeze_panes = "A3"
    ws5.auto_filter.ref = "A1:C1"

    # Chart: horizontales Balkendiagramm Root Cause
    from openpyxl.chart import BarChart, Reference
    n_rc = len(rc_counter)
    rc_chart = BarChart()
    rc_chart.type      = "bar"
    rc_chart.grouping  = "clustered"
    rc_chart.title     = f"Root Cause – abgeschlossene Bugs (n={total_with_rc})"
    rc_chart.x_axis.title = "Anzahl"
    rc_chart.y_axis.title = "Root Cause"
    rc_chart.style     = 10
    rc_chart.width     = 24
    rc_chart.height    = max(10, n_rc * 1.2)

    data_ref = Reference(ws5, min_col=2, max_col=2, min_row=2, max_row=2 + n_rc)
    cats_ref = Reference(ws5, min_col=1, max_col=1, min_row=3, max_row=2 + n_rc)
    rc_chart.add_data(data_ref, titles_from_data=True)
    rc_chart.set_categories(cats_ref)
    rc_chart.series[0].graphicalProperties.solidFill = "1F4E79"
    rc_chart.series[0].graphicalProperties.line.solidFill = "1F4E79"

    ws5.add_chart(rc_chart, "E3")

    # ── TAB 6: KOMPONENTEN ────────────────────────────────────────────────
    ws_comp = wb.create_sheet("Komponenten")

    total_bugs = len(results)

    # Komponentenfrequenz aufbauen:
    # comp_mentions  = wie oft taucht die Komponente in irgendeinem Bug auf
    # comp_defects   = wie viele (eindeutige) Bugs betreffen diese Komponente
    comp_mentions: dict[str, int] = {}
    comp_defects:  dict[str, set] = {}

    for r in results:
        if not r["components"]:
            continue
        for comp in [c.strip() for c in r["components"].split(",") if c.strip()]:
            comp_mentions[comp] = comp_mentions.get(comp, 0) + 1
            comp_defects.setdefault(comp, set()).add(r["key"])

    # Für Ø Komponenten pro Defect: Bugs mit mind. 1 Komponente
    bugs_with_comp   = sum(1 for r in results if r["components"])
    total_comp_mentions = sum(comp_mentions.values())
    avg_comp_per_bug = round(total_comp_mentions / bugs_with_comp, 2) if bugs_with_comp else 0

    # Sortiert nach Defect-Anzahl absteigend
    sorted_comps = sorted(comp_defects.keys(), key=lambda c: -len(comp_defects[c]))
    n_comp = len(sorted_comps)

    # Header
    comp_headers = ["Komponente", "Nennungen", "Betroffene Defects", "Anteil Defects %"]
    for ci, h in enumerate(comp_headers, 1):
        hdr(ws_comp, 1, ci, h)

    # Zeile 2: Info
    info_comp = (
        f"Basis: {total_bugs} Bugs gesamt  |  "
        f"{bugs_with_comp} mit Komponente  |  "
        f"Ø {avg_comp_per_bug} Komponenten pro Defect"
    )
    ws_comp.cell(row=2, column=1, value=info_comp)
    ws_comp.cell(row=2, column=1).font = Font(name="Arial", italic=True, size=9, color="595959")
    ws_comp.merge_cells("A2:D2")

    ri_c = 3
    for comp in sorted_comps:
        n_def  = len(comp_defects[comp])
        n_men  = comp_mentions[comp]
        pct    = n_def / total_bugs if total_bugs else 0
        f      = ALT_FILL if ri_c % 2 == 0 else None
        cel(ws_comp, ri_c, 1, comp,  fill=f)
        cel(ws_comp, ri_c, 2, n_men, fill=f, center=True)
        cel(ws_comp, ri_c, 3, n_def, fill=f, center=True)
        ws_comp.cell(row=ri_c, column=4).value         = pct
        ws_comp.cell(row=ri_c, column=4).number_format = "0.0%"
        ws_comp.cell(row=ri_c, column=4).alignment     = CENTER
        ws_comp.cell(row=ri_c, column=4).font          = BODY_FONT
        if f:
            ws_comp.cell(row=ri_c, column=4).fill = f
        ri_c += 1

    # Summenzeile
    ws_comp.cell(row=ri_c, column=1, value="Gesamt (Bugs mit Komponente)").font = BOLD_FONT
    ws_comp.cell(row=ri_c, column=2, value=total_comp_mentions).font            = BOLD_FONT
    ws_comp.cell(row=ri_c, column=3, value=bugs_with_comp).font                 = BOLD_FONT

    widths(ws_comp, {"A": 35, "B": 14, "C": 20, "D": 18})
    ws_comp.freeze_panes = "A3"
    ws_comp.auto_filter.ref = "A1:D1"

    # Chart: horizontaler Balken – Betroffene Defects je Komponente
    comp_chart = BarChart()
    comp_chart.type          = "bar"
    comp_chart.grouping      = "clustered"
    comp_chart.title         = f"Betroffene Defects je Komponente (n={total_bugs} Bugs gesamt)"
    comp_chart.x_axis.title  = "Anzahl Defects"
    comp_chart.y_axis.title  = "Komponente"
    comp_chart.style         = 10
    comp_chart.width         = 24
    comp_chart.height        = max(10, n_comp * 1.2)

    # Spalte C (Betroffene Defects) als Datenserie
    data_ref_c = Reference(ws_comp, min_col=3, max_col=3, min_row=2, max_row=2 + n_comp)
    cats_ref_c = Reference(ws_comp, min_col=1, max_col=1, min_row=3, max_row=2 + n_comp)
    comp_chart.add_data(data_ref_c, titles_from_data=True)
    comp_chart.set_categories(cats_ref_c)
    comp_chart.series[0].graphicalProperties.solidFill = "2E75B6"
    comp_chart.series[0].graphicalProperties.line.solidFill = "2E75B6"

    ws_comp.add_chart(comp_chart, "F3")

    # ── TAB 7: CHARTS ─────────────────────────────────────────────────────
    ws6 = wb.create_sheet("Charts")
    ws6["A1"].value = "Grafische Auswertungen"
    ws6["A1"].font  = Font(name="Arial", bold=True, size=14, color="1F4E79")

    teams = sorted({r["team_group"] for r in results})
    prios = ["Critical", "High", "Medium", "Low"]
    cats  = ["BLOCKED", "Development", "Test", "Ready For Deployment"]
    n_teams = len(teams)

    # ── Hilfsdaten: Retest je Team × Prio (Zeilen 3..3+n_teams, Spalten A..E) ──
    # Layout:
    #   Zeile 3:           Team | Critical | High | Medium | Low
    #   Zeile 4..3+n:      <team> | <cnt> | ...
    D_ROW1 = 3   # Startzeile Retest-Daten
    ws6.cell(row=D_ROW1, column=1, value="Team")
    for pi, p in enumerate(prios):
        ws6.cell(row=D_ROW1, column=2+pi, value=p)
    for ti, t in enumerate(teams):
        ws6.cell(row=D_ROW1+1+ti, column=1, value=t)
        for pi, p in enumerate(prios):
            ws6.cell(row=D_ROW1+1+ti, column=2+pi,
                     value=sum(r["retest_count"] for r in results
                               if r["team_group"] == t and r["priority"] == p))

    # ── Hilfsdaten: Liegezeit je Team × Kategorie ──────────────────────────
    # Layout (nach Retest-Block + 2 Leerzeilen):
    #   Zeile D_ROW2:          Team | BLOCKED | Development | Test | Ready For Deployment
    #   Zeile D_ROW2+1..:      <team> | <avg_h> | ...
    D_ROW2 = D_ROW1 + n_teams + 3
    ws6.cell(row=D_ROW2, column=1, value="Team")
    for ci2, c in enumerate(cats):
        ws6.cell(row=D_ROW2, column=2+ci2, value=c)
    for ti, t in enumerate(teams):
        ws6.cell(row=D_ROW2+1+ti, column=1, value=t)
        for ci2, c in enumerate(cats):
            keys_c: set = set(); sec_c = 0.0
            for r in results:
                if r["team_group"] != t:
                    continue
                for seg in r["segments"]:
                    if STATUS_CATEGORY_MAPPING.get(seg["status"], "Sonstige") == c and seg["duration_sec"] > 0:
                        keys_c.add(r["key"]); sec_c += seg["duration_sec"]
            ws6.cell(row=D_ROW2+1+ti, column=2+ci2,
                     value=round(sec_c / 3600 / len(keys_c), 2) if keys_c else 0)

    # ── Chart 1: Retest-Loops je Team (gestapelter Balken nach Prio) ──────
    c1 = BarChart()
    c1.type      = "col"          # vertikale Balken
    c1.grouping  = "stacked"
    c1.overlap   = 100
    c1.title     = "Retest-Loops je Team (nach Priorität)"
    c1.y_axis.title = "Retest-Anzahl (Summe)"
    c1.x_axis.title = "Team"
    c1.style     = 10
    c1.width     = 20
    c1.height    = 14

    # Kategorien (Team-Namen) – EINMALIG setzen
    cats_ref1 = Reference(ws6,
                          min_col=1, max_col=1,
                          min_row=D_ROW1+1, max_row=D_ROW1+n_teams)
    c1.set_categories(cats_ref1)

    prio_hex = ["C00000", "FF0000", "ED7D31", "70AD47"]
    for pi in range(len(prios)):
        data_ref = Reference(ws6,
                             min_col=2+pi, max_col=2+pi,
                             min_row=D_ROW1,            # Header-Zeile = Serienname
                             max_row=D_ROW1+n_teams)
        c1.add_data(data_ref, titles_from_data=True)
        c1.series[pi].graphicalProperties.solidFill = prio_hex[pi]
        c1.series[pi].graphicalProperties.line.solidFill = prio_hex[pi]

    ws6.add_chart(c1, "A" + str(D_ROW2 + n_teams + 4))   # unter Liegezeit-Daten

    # ── Chart 2: ∅ Liegezeit je Team & Phase (gruppierter Balken) ─────────
    c2 = BarChart()
    c2.type      = "col"
    c2.grouping  = "clustered"
    c2.title     = "∅ Liegezeit je Team und Phase (Stunden pro Defect)"
    c2.y_axis.title = "∅ Stunden pro Defect"
    c2.x_axis.title = "Team"
    c2.style     = 10
    c2.width     = 20
    c2.height    = 14

    cats_ref2 = Reference(ws6,
                          min_col=1, max_col=1,
                          min_row=D_ROW2+1, max_row=D_ROW2+n_teams)
    c2.set_categories(cats_ref2)

    cat_hex = ["C00000", "2E75B6", "70AD47", "ED7D31"]
    for ci2 in range(len(cats)):
        data_ref2 = Reference(ws6,
                              min_col=2+ci2, max_col=2+ci2,
                              min_row=D_ROW2,             # Header-Zeile = Serienname
                              max_row=D_ROW2+n_teams)
        c2.add_data(data_ref2, titles_from_data=True)
        c2.series[ci2].graphicalProperties.solidFill = cat_hex[ci2]
        c2.series[ci2].graphicalProperties.line.solidFill = cat_hex[ci2]

    # Chart 2 direkt unter Chart 1 platzieren (Chart 1 startet bei Zeile D_ROW2+n+4, Höhe ≈ 28 Zeilen)
    c2_row = D_ROW2 + n_teams + 4 + 30
    ws6.add_chart(c2, "A" + str(c2_row))

    wb.save(path)
    print(f"[Export Excel] Gespeichert: {path}")
    print(f"  Tabs: Defects | Statuswechsel | Retest-Matrix | Liegezeiten | Root Cause | Komponenten | Charts")


def print_retest_matrix(results: list[dict]) -> None:
    """
    Kompakte Matrixdarstellung:
    Defect-ID | Aggregierter Team-Name | Priorität | Retest-Anzahl

    Sortierung: Team-Gruppe → Priorität (Critical > High > Medium > Low) → Retest absteigend
    """
    PRIO_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

    sorted_results = sorted(
        results,
        key=lambda r: (
            r["team_group"],
            PRIO_ORDER.get(r["priority"], 99),
            -r["retest_count"],
        ),
    )

    # Maximale Retest-Anzahl für die Balkenbreite
    max_retest = max((r["retest_count"] for r in results), default=0)
    bar_scale  = 20 / max(max_retest, 1)   # max. 20 Zeichen Balkenbreite

    print("\n")
    print("=" * 75)
    print("RETEST-MATRIX – Projekt Fontus / Bugs")
    print("=" * 75)
    print(f"{'DEFECT-ID':<16} {'TEAM (AGG.)':<20} {'PRIO':<10} {'ABG.':>5} {'RETESTS':>7}  VERLAUF")
    print("-" * 75)

    prev_group = None
    for r in sorted_results:
        # Trennlinie bei neuem Team
        if r["team_group"] != prev_group:
            if prev_group is not None:
                print()
            prev_group = r["team_group"]

        bar    = "█" * int(r["retest_count"] * bar_scale)
        count  = f"{r['retest_count']:>3}"
        closed = {"Ja": "✓", "reject": "✗", "Nein": "⏳"}.get(r["is_closed"], r["is_closed"])
        print(f"{r['key']:<16} {r['team_group']:<20} {r['priority']:<10} {closed:>5} {count:>7}  {bar}")

    print("-" * 75)
    total_retests = sum(r["retest_count"] for r in results)
    bugs_with_retest = sum(1 for r in results if r["retest_count"] > 0)
    print(f"  Gesamt: {len(results)} Bugs  |  davon mit Retest: {bugs_with_retest}  |  Retest-Summe: {total_retests}")
    print("=" * 75)


def print_category_dwell_report(results: list[dict]) -> None:
    """
    Aggregierte Liegezeit-Auswertung nach:
      Aggregierter Team-Name × Kategorie × Priorität

    Zeigt: Anzahl Defects, Gesamtliegezeit, Durchschnittliche Liegezeit pro Defect.
    Statuswerte die nicht im STATUS_CATEGORY_MAPPING sind werden als 'Sonstige' gebucht.

    Liegezeit = Summe aller Segment-Dauern (bereinigt, ohne 0-Sekunden-Segmente)
    die dem jeweiligen Status zugeordnet sind.
    """
    from collections import defaultdict

    PRIO_ORDER = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}

    # Aggregations-Struktur:
    # key = (team_group, category, priority, is_closed)
    # value = {"total_sec": float, "defect_keys": set}
    agg: dict[tuple, dict] = defaultdict(lambda: {"total_sec": 0.0, "defect_keys": set()})

    for r in results:
        group    = r["team_group"]
        priority = r["priority"]
        closed   = r["is_closed"]

        for seg in r["segments"]:
            if seg["duration_sec"] <= 0:
                continue
            category = STATUS_CATEGORY_MAPPING.get(seg["status"], "Sonstige")
            key      = (group, category, priority, closed)
            agg[key]["total_sec"]    += seg["duration_sec"]
            agg[key]["defect_keys"].add(r["key"])

    # Sortieren: Team → Kategorie → Prio → Ja zuerst, dann Nein, dann reject
    CLOSED_ORDER   = {"Ja": 0, "Nein": 1, "reject": 2}
    CATEGORY_ORDER = {"BLOCKED": 0, "Development": 1, "Test": 2, "Ready For Deployment": 3, "Sonstige": 99}
    sorted_keys = sorted(
        agg.keys(),
        key=lambda k: (k[0], CATEGORY_ORDER.get(k[1], 99), PRIO_ORDER.get(k[2], 99), CLOSED_ORDER.get(k[3], 9)),
    )

    print("\n")
    print("=" * 100)
    print("LIEGEZEIT-AUSWERTUNG NACH KATEGORIE – Projekt Fontus / Bugs")
    print("=" * 100)
    print(f"{'TEAM (AGG.)':<22} {'KATEGORIE':<22} {'PRIO':<10} {'ABG.':>8} {'DEFECTS':>7}  {'∅ LIEGEZEIT':<16} {'GESAMT'}")
    print("-" * 100)

    prev_group = None
    for k in sorted_keys:
        group, category, priority, closed = k
        data        = agg[k]
        defect_cnt  = len(data["defect_keys"])
        total_sec   = data["total_sec"]
        avg_sec     = total_sec / defect_cnt if defect_cnt else 0

        if group != prev_group:
            if prev_group is not None:
                print()
            prev_group = group

        print(
            f"{group:<22} {category:<22} {priority:<10} {closed:>8} {defect_cnt:>7}  "
            f"{fmt_duration(avg_sec):<16} {fmt_duration(total_sec)}"
        )

    print("-" * 100)

    # Gesamtzeile je Kategorie × Abgeschlossen-Kennung
    cat_totals: dict[tuple, dict] = defaultdict(lambda: {"total_sec": 0.0, "defect_keys": set()})
    for k, data in agg.items():
        cat_totals[(k[1], k[3])]["total_sec"]    += data["total_sec"]
        cat_totals[(k[1], k[3])]["defect_keys"] |= data["defect_keys"]

    print()
    print(f"  {'KATEGORIE':<22} {'ABG.':>8} {'DEFECTS':>7}  {'∅ LIEGEZEIT':<16} {'GESAMT'}")
    print(f"  {'-'*70}")
    for cat_closed in sorted(cat_totals.keys(), key=lambda c: (CATEGORY_ORDER.get(c[0], 99), CLOSED_ORDER.get(c[1], 9))):
        cat, closed = cat_closed
        d   = cat_totals[cat_closed]
        cnt = len(d["defect_keys"])
        avg = d["total_sec"] / cnt if cnt else 0
        print(f"  {cat:<22} {closed:>8} {cnt:>7}  {fmt_duration(avg):<16} {fmt_duration(d['total_sec'])}")
    print("=" * 100)


# ===========================================================================
# DATENBANK-EXPORT (SQLite)
# ===========================================================================

def export_sqlite(results: list[dict], run_ts: str,
                  db_path: str = None) -> None:
    """
    Speichert alle Auswertungsergebnisse in eine SQLite-Datenbank.

    Tabellen:
      runs             – Ein Eintrag pro Analyselauf (Primärschlüssel: run_ts)
      defects          – Stammdaten je Bug × Lauf
      segments         – Status-Segmente je Bug × Lauf
      retest_matrix    – Retest-Anzahl je Bug × Lauf
      dwell_times      – Aggregierte Liegezeiten je Team × Kategorie × Prio × Abg. × Lauf
      root_cause       – Root-Cause-Häufigkeiten je Lauf
      components       – Komponenten-Häufigkeiten je Lauf

    Schlüsselkonzept:
      run_ts (TEXT, Format: YYYYMMDD_HHMM) ist der Lauf-Zeitstempel.
      Er verbindet alle Tabellen und ermöglicht Trendvergleiche in Power BI.
      Bei erneutem Einfügen desselben run_ts werden die alten Daten
      automatisch gelöscht und neu geschrieben (REPLACE-Strategie).
    """
    import sqlite3
    from collections import defaultdict, Counter

    if db_path is None:
        db_path = os.getenv("JIRA_DB_PATH", "fontus_analysis.db")

    print(f"\n[Export SQLite] Verbinde mit: {db_path}")
    con = sqlite3.connect(db_path)
    cur = con.cursor()

    # ------------------------------------------------------------------
    # Schema erstellen (falls noch nicht vorhanden)
    # ------------------------------------------------------------------
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS runs (
            run_ts       TEXT PRIMARY KEY,
            created_at   TEXT NOT NULL,
            total_bugs   INTEGER,
            project      TEXT
        );

        CREATE TABLE IF NOT EXISTS defects (
            run_ts       TEXT  NOT NULL,
            key          TEXT  NOT NULL,
            summary      TEXT,
            team         TEXT,
            team_group   TEXT,
            status       TEXT,
            is_closed    TEXT,
            priority     TEXT,
            source       TEXT,
            root_cause   TEXT,
            test_level   TEXT,
            components   TEXT,
            releases     TEXT,
            retest_count INTEGER,
            PRIMARY KEY (run_ts, key)
        );

        CREATE TABLE IF NOT EXISTS segments (
            run_ts       TEXT    NOT NULL,
            key          TEXT    NOT NULL,
            seg_index    INTEGER NOT NULL,
            status       TEXT,
            category     TEXT,
            entered_at   TEXT,
            left_at      TEXT,
            duration_sec REAL,
            duration_fmt TEXT,
            PRIMARY KEY (run_ts, key, seg_index)
        );

        CREATE TABLE IF NOT EXISTS retest_matrix (
            run_ts       TEXT    NOT NULL,
            key          TEXT    NOT NULL,
            team_group   TEXT,
            priority     TEXT,
            is_closed    TEXT,
            retest_count INTEGER,
            PRIMARY KEY (run_ts, key)
        );

        CREATE TABLE IF NOT EXISTS dwell_times (
            run_ts       TEXT    NOT NULL,
            team_group   TEXT    NOT NULL,
            category     TEXT    NOT NULL,
            priority     TEXT    NOT NULL,
            is_closed    TEXT    NOT NULL,
            defect_count INTEGER,
            total_sec    REAL,
            avg_sec      REAL,
            PRIMARY KEY (run_ts, team_group, category, priority, is_closed)
        );

        CREATE TABLE IF NOT EXISTS root_cause (
            run_ts          TEXT    NOT NULL,
            root_cause      TEXT    NOT NULL,
            count           INTEGER,
            pct             REAL,
            total_with_rc   INTEGER,
            total_closed    INTEGER,
            PRIMARY KEY (run_ts, root_cause)
        );

        CREATE TABLE IF NOT EXISTS components (
            run_ts          TEXT    NOT NULL,
            component       TEXT    NOT NULL,
            mentions        INTEGER,
            defect_count    INTEGER,
            pct_of_total    REAL,
            total_bugs      INTEGER,
            PRIMARY KEY (run_ts, component)
        );
    """)
    con.commit()

    # ------------------------------------------------------------------
    # Alten Lauf löschen falls run_ts bereits existiert (Idempotenz)
    # ------------------------------------------------------------------
    for table in ("runs", "defects", "segments", "retest_matrix",
                  "dwell_times", "root_cause", "components"):
        cur.execute(f"DELETE FROM {table} WHERE run_ts = ?", (run_ts,))
    con.commit()
    print(f"  Lauf-ID: {run_ts}  (ältere Daten mit gleichem Timestamp überschrieben)")

    # ------------------------------------------------------------------
    # runs
    # ------------------------------------------------------------------
    cur.execute(
        "INSERT INTO runs (run_ts, created_at, total_bugs, project) VALUES (?,?,?,?)",
        (run_ts, datetime.now().isoformat(timespec="seconds"), len(results), PROJECT_KEY)
    )

    # ------------------------------------------------------------------
    # defects
    # ------------------------------------------------------------------
    cur.executemany(
        """INSERT INTO defects
           (run_ts,key,summary,team,team_group,status,is_closed,priority,
            source,root_cause,test_level,components,releases,retest_count)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        [
            (run_ts, r["key"], r["summary"], r["team"], r["team_group"],
             r["status"], r["is_closed"], r["priority"], r["source"],
             r["root_cause"], r["test_level"], r["components"],
             r["releases"], r["retest_count"])
            for r in results
        ]
    )

    # ------------------------------------------------------------------
    # segments
    # ------------------------------------------------------------------
    seg_rows = []
    for r in results:
        for idx, seg in enumerate(r["segments"]):
            cat = STATUS_CATEGORY_MAPPING.get(seg["status"], "Sonstige")
            seg_rows.append((
                run_ts, r["key"], idx,
                seg["status"], cat,
                seg["entered_at"], seg["left_at"],
                seg["duration_sec"], seg["duration_fmt"],
            ))
    cur.executemany(
        """INSERT INTO segments
           (run_ts,key,seg_index,status,category,entered_at,left_at,
            duration_sec,duration_fmt)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        seg_rows
    )

    # ------------------------------------------------------------------
    # retest_matrix
    # ------------------------------------------------------------------
    cur.executemany(
        """INSERT INTO retest_matrix
           (run_ts,key,team_group,priority,is_closed,retest_count)
           VALUES (?,?,?,?,?,?)""",
        [
            (run_ts, r["key"], r["team_group"], r["priority"],
             r["is_closed"], r["retest_count"])
            for r in results
        ]
    )

    # ------------------------------------------------------------------
    # dwell_times  (aggregiert wie im Excel-Tab Liegezeiten)
    # ------------------------------------------------------------------
    agg_dwell: dict[tuple, dict] = defaultdict(
        lambda: {"total_sec": 0.0, "defect_keys": set()}
    )
    for r in results:
        for seg in r["segments"]:
            if seg["duration_sec"] <= 0:
                continue
            cat = STATUS_CATEGORY_MAPPING.get(seg["status"], "Sonstige")
            key = (r["team_group"], cat, r["priority"], r["is_closed"])
            agg_dwell[key]["total_sec"]    += seg["duration_sec"]
            agg_dwell[key]["defect_keys"].add(r["key"])

    dwell_rows = []
    for (grp, cat, prio, closed), data in agg_dwell.items():
        cnt     = len(data["defect_keys"])
        total_s = data["total_sec"]
        avg_s   = total_s / cnt if cnt else 0.0
        dwell_rows.append((run_ts, grp, cat, prio, closed, cnt, total_s, avg_s))

    cur.executemany(
        """INSERT INTO dwell_times
           (run_ts,team_group,category,priority,is_closed,
            defect_count,total_sec,avg_sec)
           VALUES (?,?,?,?,?,?,?,?)""",
        dwell_rows
    )

    # ------------------------------------------------------------------
    # root_cause  (nur is_closed='Ja', ohne leere RC-Felder)
    # ------------------------------------------------------------------
    closed_with_rc = [r for r in results
                      if r["is_closed"] == "Ja" and r["root_cause"]]
    total_closed   = sum(1 for r in results if r["is_closed"] == "Ja")
    total_with_rc  = len(closed_with_rc)
    rc_cnt: Counter = Counter(r["root_cause"] for r in closed_with_rc)

    cur.executemany(
        """INSERT INTO root_cause
           (run_ts,root_cause,count,pct,total_with_rc,total_closed)
           VALUES (?,?,?,?,?,?)""",
        [
            (run_ts, rc, cnt,
             round(cnt / total_with_rc * 100, 2) if total_with_rc else 0,
             total_with_rc, total_closed)
            for rc, cnt in rc_cnt.most_common()
        ]
    )

    # ------------------------------------------------------------------
    # components
    # ------------------------------------------------------------------
    total_bugs     = len(results)
    comp_mentions: dict[str, int] = {}
    comp_defects:  dict[str, set] = {}
    for r in results:
        if not r["components"]:
            continue
        for comp in [c.strip() for c in r["components"].split(",") if c.strip()]:
            comp_mentions[comp] = comp_mentions.get(comp, 0) + 1
            comp_defects.setdefault(comp, set()).add(r["key"])

    cur.executemany(
        """INSERT INTO components
           (run_ts,component,mentions,defect_count,pct_of_total,total_bugs)
           VALUES (?,?,?,?,?,?)""",
        [
            (run_ts, comp,
             comp_mentions[comp],
             len(comp_defects[comp]),
             round(len(comp_defects[comp]) / total_bugs * 100, 2) if total_bugs else 0,
             total_bugs)
            for comp in sorted(comp_defects, key=lambda c: -len(comp_defects[c]))
        ]
    )

    con.commit()
    con.close()

    # Zeilenzählung für Bestätigung
    con2 = sqlite3.connect(db_path)
    totals = {}
    for tbl in ("runs", "defects", "segments", "retest_matrix",
                "dwell_times", "root_cause", "components"):
        totals[tbl] = con2.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
    con2.close()

    print(f"[Export SQLite] Gespeichert: {db_path}")
    for tbl, n in totals.items():
        print(f"  {tbl:<20} {n:>6} Zeilen gesamt")


# ===========================================================================
# MAIN
# ===========================================================================

if __name__ == "__main__":
    # --- Felderkennung (einmalig) ---
    field_id = FEATURE_TEAM_FIELD or discover_feature_team_field("feature team")

    if field_id and not FEATURE_TEAM_FIELD:
        print(f"  → Tipp: Trage in .env ein:  JIRA_FEATURE_TEAM_FIELD={field_id}")
        print(f"           um die Felderkennung beim nächsten Start zu überspringen.\n")

    # --- Schritt 1 ---
    bugs = fetch_bugs(feature_team_field=field_id or "", feature_teams=FEATURE_TEAMS, exclude_labels=EXCLUDE_LABELS, status_filter=STATUS_FILTER)
    print_bug_summary(bugs, field_id=field_id or "")

    # --- Schritt 2 ---
    results = analyse_transitions(bugs, field_id=field_id or "")
    print_transition_report(results)

    # --- Retest-Matrix ---
    print_retest_matrix(results)

    # --- Kategorisierte Liegezeit-Auswertung ---
    print_category_dwell_report(results)

    # --- JSON-Export ---
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    export_json(results, f"fontus_transitions_{ts}.json")

    # --- Excel-Export ---
    export_excel(results, f"fontus_analysis_{ts}.xlsx")

    # --- SQLite-Export ---
    export_sqlite(results, run_ts=ts, db_path=DB_PATH)
